import pandas as pd
import os
import matplotlib.pyplot as plt
import numpy as np
from matplotlib.widgets import Cursor
from mpl_toolkits.mplot3d import proj3d
from POINT import Point
from PART import Part
from REGISTERED_POINT import RegisteredPoint
from FILE_INFO import FileInfo
import tkinter as tk
from tkinter import ttk, font
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg, NavigationToolbar2Tk
from matplotlib.backend_bases import cursors
from openpyxl import load_workbook
from tkintertable import TableCanvas, TableModel


class App:
    def __init__(self):
        self.draw_start_gui()
        # Initial Settings
        self.a = 0
        self.c = 0
        self.state = 'MENU'
        self.plane_type = '-'  # VARIABLE A ESCOGER (v900/v1000/RPB)
        self.configuration = '-'  # VARIABLE A ESCOGER (Part/Quadrant)
        self.selected_quadrant = '-'  # VARIABLE A ESCOGER (Upper/Lower/Right/Left)
        self.selected_part = '-'  # VARIABLE A ESCOGER (nombre de la parte)
        self.matchers = ['-FC-', '-FCFC-', '-FCFCFC-', '-FCFCFCFC-', '-FCTI-', '-FCTIAL-', '-FCTIFC-', '-FCTIFCAL-',
                         '-FCTIALAL-']  # Variable a escoger
        self.selected_tool_types_1 = [503, 534, 535]  # VARIABLEA ESCOGER
        self.selected_tool_types = ['ADH', 'TDRILL']  # VARIABLE A ESCOGER
        self.registered_point_table_id = '-'  # Contador de id necesario para crear la tabla de la GUI

        while True:  # Main loop
            # Run back_end code
            self.update_code()
            # print(self.state, self.plane_type, self.configuration, self.selected_quadrant, self.selected_part)

    def update_code(self):
        # Intial settings
        if self.state == 'MENU':
            self.start_screen.update_idletasks()
            self.start_screen.update()

        if self.state == 'LOADING':
            print('loading')
            # Abrir y Leer archivo .xlsx
            self.main_root = os.path.dirname(os.path.abspath(__file__))
            if self.plane_type == 'v900':
                self.xlsx_name = 'v900.xlsx'
            if self.plane_type == 'v1000':
                self.xlsx_name = 'v1000.xlsx'
            if self.plane_type == 'RPB':
                self.xlsx_name = 'RPB.xlsx'
            self.file_root = os.path.join(self.main_root, self.xlsx_name)
            self.df = pd.read_excel(self.file_root)

            # Leer cada columna y crear una lista de cada una
            self.ids = self.df['ID']
            self.process_feature_names = self.df['Process Feature Name']
            self.diameters = self.df['Diameter']
            self.xes = self.df['Xe']
            self.yes = self.df['Ye']
            self.zes = self.df['Ze']
            self.tool_1s = self.df['Tool 1']
            self.tools = self.df['Tdrill/adh/230/ninguno']
            self.quadrants = self.df['CUADRANTE TEORICO']

            self.all_points = []  # Todos lo puntos que hay en el excel
            # Crear una lista con todos los objetos point
            self.dict_point_parts = self.get_point_part_names()
            for i in range(len(self.ids)):
                id = self.ids[i]
                process_f_name = self.process_feature_names[i]
                diameter = self.diameters[i]
                xe = self.xes[i]
                ye = self.yes[i]
                ze = self.zes[i]
                tool_1 = self.tool_1s[i]
                tool = self.tools[i]
                quadrant = self.quadrants[i]
                try:
                    part_names = self.dict_point_parts[id]
                except:
                    part_names = ['-']
                point = Point(id, process_f_name, diameter, xe, ye, ze, tool_1,
                              tool, quadrant, part_names)
                self.all_points.append(point)

            # Crear todas las partes
            self.part_names, self.all_parts = self.get_all_parts()
            # Leer registro de calidad de puntos
            self.registered_points = self.read_record()

            print('Se han creado {} puntos'.format(len(self.all_points)))
            print('Se han creado {} partes'.format(len(self.all_parts)))
            self.state = 'SETTINGS'

        if self.state == 'SETTINGS':
            if self.a == 0:
                self.start_screen.destroy()
                self.draw_main_gui()
                self.frame2.tkraise()
            self.main.update()
            self.main.update_idletasks()
            self.a += 1

    # Main displays functions
    def draw_start_gui(self):
        # Start TKinter
        self.start_screen = tk.Tk()
        # Create a style
        style = ttk.Style(self.start_screen)
        self.start_screen.tk.call('source', 'azure.tcl')
        style.theme_use('azure')

        # Screen settings
        windowWidth = 800
        windowHeight = 800
        screenWidth = self.start_screen.winfo_screenwidth()
        screenHeight = self.start_screen.winfo_screenheight()
        xCordinate = int((screenWidth / 2) - (windowWidth / 2))
        yCordinate = int((screenHeight / 2) - (windowHeight / 2))
        self.start_screen.geometry("{}x{}+{}+{}".format(windowWidth, windowHeight, xCordinate, yCordinate))

        # Creating a Font object of "TkDefaultFont"
        self.defaultFont = font.nametofont("TkDefaultFont")

        # Overriding default-font with custom settings
        # i.e changing font-family, size and weight
        self.defaultFont.configure(family=None,
                                   size=15,
                                   weight=font.NORMAL)
        self.select_record_window()

    def draw_main_gui(self):
        self.main = tk.Tk()
        # Create a style
        style = ttk.Style(self.main)
        self.main.tk.call('source', 'azure.tcl')
        style.theme_use('azure')
        self.main.state('zoomed')  # Fit the whole screen
        style.configure('TNotebook.Tab', font=('Arial', '11', 'normal'))
        style.configure('Treeview', rowheight=30, font=('Arial', '15', 'normal'))

        # Cambiar letra
        self.defaultFont = font.nametofont("TkDefaultFont")
        self.defaultFont.configure(family="Arial",
                                   size=20,
                                   weight=font.NORMAL)
        # Menu de configuracion de parte
        self.part_window()

        # Menu de configuracion de cuadrante
        self.quadrant_window()

    # GUI Functions
    def select_record_window(self):
        # Main Menu Window
        self.frame1 = ttk.Frame(self.start_screen)
        self.frame1.place(relwidth=1, relheight=1)
        n_cols = 2
        n_rows = 2
        weight = 1
        for i in range(0, n_cols):
            tk.Grid.columnconfigure(self.frame1, i, weight=weight)

        for i in range(0, n_rows):
            if i in [0]:
                tk.Grid.rowconfigure(self.frame1, i, weight=0)
            else:
                tk.Grid.rowconfigure(self.frame1, i, weight=weight)

        # Titulo
        self.title1 = tk.Label(self.frame1, text='Digitalización de Calidad', font=('Arial', '24'))
        self.title1.grid(row=0, column=0, sticky='w', padx=10)
        # Lista de archivos
        self.files_frame = tk.LabelFrame(self.frame1, text='Selecciona un registro')
        self.files_frame.grid(row=1, column=0, columnspan=2, sticky='ewns', padx=10, pady=10)
        for i in range(0, 2):
            tk.Grid.columnconfigure(self.files_frame, i, weight=1)
        for i in range(0, 2):
            if i == 0:
                tk.Grid.rowconfigure(self.files_frame, i, weight=0)
            else:
                tk.Grid.rowconfigure(self.files_frame, i, weight=1)
        # Lista de partes
        self.list_frame = tk.Frame(self.files_frame)
        self.list_frame.grid(row=1, column=0, columnspan=3, sticky='wens')
        self.list1 = tk.Listbox(self.list_frame, bd=0, bg='#CBCBCB', selectbackground='#007fff')

        self.files_list = []
        self.all_files_info = self.read_file_names()
        for file in self.all_files_info:
            self.files_list.append(file.name)
        self.update_parts_list(self.files_list)
        self.list1.pack(side='left', fill='both', expand=True)
        # Scrollbar
        self.scroll = tk.Scrollbar(self.list_frame, orient='vertical', width=25)
        self.scroll.config(command=self.list1.yview)
        self.scroll.pack(side='right', fill='y')
        self.list1.config(yscrollcommand=self.scroll.set)
        # Boton de agregar un nuevo registro
        self.button5 = ttk.Button(self.files_frame, text='Nuevo avión', style='AccentButton',
                                  command=lambda: [self.new_record_window()])
        self.button5.grid(row=0, column=1, sticky='s', padx=10, pady=(0, 5))
        # Boton de Start
        self.button6 = ttk.Button(self.files_frame, text='Iniciar avión', style='AccentButton',
                                  command=lambda: [self.set_active_record(), self.change_state('LOADING')])
        self.button6.grid(row=0, column=0, sticky='s', padx=10, pady=(0, 5))

    def new_record_window(self):
        # Start TKinter
        self.create_record_screen = tk.Tk()
        # Create a style
        style = ttk.Style(self.create_record_screen)
        self.create_record_screen.tk.call('source', 'azure.tcl')
        style.theme_use('azure')

        # Screen settings
        windowWidth = 350
        windowHeight = 180
        screenWidth = self.create_record_screen.winfo_screenwidth()
        screenHeight = self.create_record_screen.winfo_screenheight()
        xCordinate = int((screenWidth / 2) - (windowWidth / 2))
        yCordinate = int((screenHeight / 2) - (windowHeight / 2))
        self.create_record_screen.geometry("{}x{}+{}+{}".format(windowWidth, windowHeight, xCordinate, yCordinate))

        # Creating a Font object of "TkDefaultFont"
        self.defaultFont = font.nametofont("TkDefaultFont")

        # Overriding default-font with custom settings
        # i.e changing font-family, size and weight
        self.defaultFont.configure(family=None,
                                   size=18,
                                   weight=font.NORMAL)
        # Frame
        self.frame3 = tk.LabelFrame(self.create_record_screen, text='CREAR UN NUEVO REGISTRO')
        self.frame3.pack(fill='both', expand=True, padx=15, pady=15)
        n_cols = 2
        n_rows = 3
        weight = 1
        for i in range(0, n_cols):
            tk.Grid.columnconfigure(self.frame3, i, weight=weight)

        for i in range(0, n_rows):
            if i in [121412]:
                tk.Grid.rowconfigure(self.frame3, i, weight=0)
            else:
                tk.Grid.rowconfigure(self.frame3, i, weight=weight)
        # Nombre del registro:
        self.label4 = tk.Label(self.frame3, text='Nombre del avión:')
        self.label4.grid(row=0, column=0)
        # Entrada de texto
        self.entry3 = tk.Entry(self.frame3)
        self.entry3.grid(row=0, column=1)
        # Texto tipo de avion
        self.label5 = tk.Label(self.frame3, text='Tipo de avion')
        self.label5.grid(row=1, column=0)
        # Tipo de avion
        self.button7 = ttk.Combobox(self.frame3, state='readonly', values=['v900', 'v1000', 'RPB'])
        self.button7.grid(row=1, column=1)
        self.button7.bind('<<ComboboxSelected>>', self.set_plane_type)
        # Boton crear registro
        # self.button8 = ttk.Button(self.frame3, text='CREAR', style='AccentButton',
        #                           command=lambda: [self.change_state('LOADING'),
        #                                            self.add_record_info(self.entry3.get(), self.plane_type),
        #                                            self.create_record_screen.destroy()])
        self.button8 = ttk.Button(self.frame3, text='CREAR', style='AccentButton',
                                  command=lambda: [self.create_record()])
        self.button8.grid(row=2, column=0, columnspan=2)

    def quadrant_window(self):
        # Marco global de Cuadrante
        self.frame2 = ttk.Frame(self.main)
        self.frame2.place(relwidth=1, relheight=1)
        n_cols = 4
        n_rows = 5
        weight = 1
        for i in range(0, n_cols):
            tk.Grid.columnconfigure(self.frame2, i, weight=weight)

        for i in range(0, n_rows):
            if i in [0, 1, 2]:
                tk.Grid.rowconfigure(self.frame2, i, weight=0)
            else:
                tk.Grid.rowconfigure(self.frame2, i, weight=weight)
        # Titulo
        title = str(self.selected_record.name) + ' - ' + str(self.plane_type)
        self.label2 = ttk.Label(self.frame2, text=title)
        self.label2.grid(row=0, column=0, sticky='w', pady=(0, 5), padx=(10, 10))
        # Separator
        self.separator = ttk.Separator(self.frame2)
        self.separator.grid(row=1, column=0, columnspan=4, sticky='we', pady=(20, 30), padx=(10, 10))
        # Button de pieza/cuadrante
        self.button3 = ttk.Combobox(self.frame2, state='readonly', values=['Part', 'Quadrant'], width=8)
        self.button3.grid(row=0, column=1, sticky='s', pady=(0, 20), padx=(10, 0))
        self.button3.current(1)
        self.button3.bind("<<ComboboxSelected>>", self.set_config2)
        # Boton de seleccionar cuadrante
        self.button4 = ttk.Combobox(self.frame2, state='readonly',
                                    values=['Selecciona un cuadrante', 'Upper', 'Lower', 'Right', 'Left'])
        self.button4.grid(row=0, column=2, sticky='s', pady=(0, 20))
        self.button4.current(0)
        self.button4.bind("<<ComboboxSelected>>", self.set_quadrant)
        # Cuadrante vacio
        self.fig = plt.figure(dpi=100, frameon=False, figsize=(5, 4))
        self.figure = self.fig.add_subplot()

        miniframe = tk.LabelFrame(self.frame2, text='Cuadrante')
        miniframe.grid(row=3, column=0, rowspan=2, columnspan=2, sticky='ewns', pady=(0, 20), padx=(10, 10))
        self.canvas_plot = FigureCanvasTkAgg(self.fig, miniframe)
        self.canvas_plot.get_tk_widget().pack(fill='both', side='top', expand=True)
        toolbar = NavigationToolbar2Tk(self.canvas_plot, miniframe).pack(fill='both', side='top', expand=True)
        # Visualizador de partes/info
        self.part_window = ttk.Notebook(self.frame2)
        self.part_window.grid(row=3, column=2, sticky='wens', columnspan=2, rowspan=2, pady=(0, 20), padx=(0, 10))
        # Parte vacia
        self.fig2 = plt.figure()
        self.ax_2 = self.fig2.add_subplot(projection='3d')
        miniframe = tk.Frame(self.part_window)
        self.part_window.add(miniframe, text='-')
        canvas_plot = FigureCanvasTkAgg(self.fig2, miniframe)
        canvas_plot.get_tk_widget().pack(fill='both', side='top', expand=True)
        # Caja de informacion vacia
        self.infobox = tk.LabelFrame(miniframe, text='Datos del punto')
        self.infobox.pack(fill='both', side='bottom', expand=True)
        self.infolabel = tk.Label(self.infobox, text='\n\n\n\n\n\n\n\n\n', font=('Arial', '14'))
        self.infolabel.pack(fill='both', expand=True)

    def part_window(self):
        # Marco global de Parte
        self.frame = ttk.Frame(self.main)
        self.frame.place(relwidth=1, relheight=1)
        n_cols = 3 + 1
        n_rows = 5
        weight = 1
        for i in range(0, n_cols):
            tk.Grid.columnconfigure(self.frame, i, weight=weight)
        for i in range(0, n_rows):
            if i in [0, 1, 2]:
                tk.Grid.rowconfigure(self.frame, i, weight=0)
            else:
                tk.Grid.rowconfigure(self.frame, i, weight=weight)
        # Titulo
        title = str(self.selected_record.name) + ' - ' + str(self.plane_type)
        self.label2 = ttk.Label(self.frame, text=title)
        self.label2.grid(row=0, column=0, sticky='w', columnspan=3, pady=(0, 0), padx=(10, 10))
        # Separator
        self.separator = ttk.Separator(self.frame)
        self.separator.grid(row=1, column=0, columnspan=3, sticky='we', pady=(20, 30), padx=(10, 10))
        # Button de pieza/cuadrante
        self.button1 = ttk.Combobox(self.frame, state='readonly', values=['Part', 'Quadrant'], width=8)
        self.button1.grid(row=0, column=2, sticky='w', pady=(0, 20), padx=(10, 0))
        self.button1.bind("<<ComboboxSelected>>", self.set_config1)
        # Boton start
        self.button2 = ttk.Button(self.frame, text='VER PARTE', style='AccentButton',
                                  command=lambda: [self.set_part(),
                                                   self.plot_part(self.all_parts, self.selected_part)])
        self.button2.grid(row=0, column=1, sticky='w', pady=(0, 20))

        # Cuadro de partes
        self.labelframe1 = tk.LabelFrame(self.frame, text='Nombre de la parte', font=('Arial', '17'))
        self.labelframe1.grid(row=3, column=0, columnspan=1, sticky='wens', padx=(5, 5), pady=(0, 10), rowspan=2)
        for i in range(0, 2):
            tk.Grid.columnconfigure(self.labelframe1, i, weight=1)
        for i in range(0, 2):
            if i == 0:
                tk.Grid.rowconfigure(self.labelframe1, i, weight=0)
            else:
                tk.Grid.rowconfigure(self.labelframe1, i, weight=1)

        # Entrada de texto
        self.entry1 = ttk.Entry(self.labelframe1, width=8)
        self.entry1.bind("<KeyRelease>", self.check_part_name)
        self.entry1.grid(row=0, column=0, sticky='we')
        # Lista de partes
        self.list_frame = tk.Frame(self.labelframe1)
        self.list_frame.grid(row=1, column=0, columnspan=2, sticky='wens')
        self.list1 = tk.Listbox(self.list_frame, bd=0, bg='#CBCBCB', selectbackground='#007fff')
        self.update_parts_list(self.part_names)
        self.update_parts_list_color()
        self.list1.pack(side='left', fill='both', expand=True)
        # Scrollbar
        self.scroll = tk.Scrollbar(self.list_frame, orient='vertical', width=25)
        self.scroll.config(command=self.list1.yview)
        self.scroll.pack(side='right', fill='y')
        self.list1.config(yscrollcommand=self.scroll.set)
        # Plot vacio
        self.fig_2 = plt.figure(figsize=(5, 4))
        self.fig_2ax_2 = self.fig_2.add_subplot(projection='3d')
        self.fig_2ax_2.set_title('Parte seleccionada')
        # Crear el widget
        miniframe = tk.LabelFrame(self.frame, text='Parte seleccionada', font=('Arial', '17'))
        miniframe.grid(row=3, column=1, sticky='ewns', pady=(0, 20), padx=(0, 10), rowspan=2)
        self.canvas_part = FigureCanvasTkAgg(self.fig_2, miniframe)
        self.canvas_part.get_tk_widget().pack(fill='both', expand=True)
        # Crear tabla de registro
        self.tableframe = tk.Label(self.frame)
        self.tableframe.grid(row=3, column=2, sticky='ewns', columnspan=1, padx=(0, 5), pady=(10, 0))
        self.table = ttk.Treeview(self.tableframe, show='headings', columns=(1, 2, 3, 4, 5))
        headers = ['Part Name', 'Quality', 'ID', 'Problem', 'Quadrant']
        for i in range(0, len(headers)):
            self.table.heading(i + 1, text=headers[i], anchor='w')
        self.table.pack(fill='both', expand=True, side='left')
        # Barra de scroll de tabla
        scrollbar = tk.Scrollbar(self.tableframe, orient='vertical', width=25)
        scrollbar.config(command=self.table.yview)
        self.table.config(yscrollcommand=scrollbar.set)
        scrollbar.pack(fill='y', side='right')
        # Agregar datos a la tabla
        self.add_table_data(self.registered_points)
        # Create item form frame
        self.entryframe = tk.LabelFrame(self.frame, text='Registrar Punto', font=('Arial', '18'))
        self.entryframe.grid(row=4, column=2, sticky='ewns', columnspan=2, pady=(0, 20), padx=(0, 10))
        for i in range(0, 4):
            tk.Grid.columnconfigure(self.entryframe, i, weight=1)
        for i in range(0, 4):
            tk.Grid.rowconfigure(self.entryframe, i, weight=1)
        # Label del nombre de la parte
        label3 = ttk.Label(self.entryframe, text='Part name:', font=('Arial', '18'))
        label3.grid(row=0, column=0, sticky='w')
        # Nombre de la parte
        self.entry4 = tk.Button(self.entryframe, text='empty', bg='white', bd=0, width=22, anchor='w',
                                font=('Arial', '18'))  # PARTE
        self.entry4.grid(row=0, column=1, sticky='wns')
        # Label de id
        label2 = ttk.Label(self.entryframe, text='ID:', font=('Arial', '18'))
        label2.grid(row=0, column=3, sticky='w')
        # Boton de id
        self.entry6 = tk.Button(self.entryframe, text='empty', bg='white', bd=0, anchor='w', font=('Arial', '18'))  # ID
        self.entry6.grid(row=0, column=3, sticky='e')
        # Label de calidad
        label = ttk.Label(self.entryframe, text='Quality:', font=('Arial', '18'))
        label.grid(row=1, column=0, sticky='w')
        # Boton de calidad
        self.entry5 = ttk.Combobox(self.entryframe, state='readonly', values=['1', '0'], width=2)
        self.entry5.bind("<<ComboboxSelected>>", self.set_quality)
        self.entry5.grid(row=1, column=1, sticky='wns', pady=(5, 5))
        # Label de Quadrante
        label4 = ttk.Label(self.entryframe, text='Quadrant:', font=('Arial', '18'))
        label4.grid(row=1, column=3, sticky='w')
        # Boton de Quadrante
        self.entry8 = tk.Button(self.entryframe, text='empty', bg='white', bd=0, anchor='w', font=('Arial', '18'))  # Quadrant
        self.entry8.grid(row=1, column=3, sticky='e')
        # Label de problem type
        label3 = ttk.Label(self.entryframe, text='Problem Type:', font=('Arial', '18'))
        label3.grid(row=2, column=0, sticky='w')
        problem_types = ['Pequeño', 'Primera', 'Segunda', 'Tercera',
                         'NC', 'Quemadura', 'Delaminacion', 'Otro']
        # Boton de problem type
        self.entry7 = ttk.Combobox(self.entryframe, values=problem_types, state='readonly', width=19)
        self.entry7.grid(row=2, column=1, sticky='wns', pady=(5, 5))
        # Boton de eliminar datos
        self.button10 = ttk.Button(self.entryframe, text='Delete Entry', command=lambda: self.delete_entry(),
                                   style='AccentButton')
        self.button10.grid(row=3, column=1, pady=(5, 5))
        # Boton de agregar datos
        self.button9 = ttk.Button(self.entryframe, text='Add', command=lambda: self.add_entry(), style='AccentButton')
        self.button9.grid(row=3, column=3, pady=(5, 5))

    def select_quadrant(self):
        self.selected_quadrant = self.q.get()
        if self.plane_type in ['v1000', 'RPB']:
            if self.selected_quadrant == 'Left':
                self.selected_quadrant = 'Izquierdo'
            elif self.selected_quadrant == 'Right':
                self.selected_quadrant = 'Derecho'
            elif self.selected_quadrant == 'Upper':
                self.selected_quadrant = 'Superior'
            elif self.selected_quadrant == 'Lower':
                self.selected_quadrant = 'Inferior'

    def set_config1(self, event):
        print(self.button1.get())
        if self.button1.get() == 'Part':
            self.configuration = 'Part'
            self.frame.tkraise()
            self.button3.current(0)
        if self.button1.get() == 'Quadrant':
            self.configuration = 'Quadrant'
            self.frame2.tkraise()
            self.button3.current(1)

    def set_quality(self, event):
        print(self.entry5.get())
        if str(self.entry5.get()) == '1':
            self.entry6.config(text=self.selected_registered_point_id)
            problem_types = ['Pequeño', 'Primera', 'Segunda', 'Tercera',
                             'NC', 'Quemadura', 'Delaminacion', 'Otro']
            # Boton de problem type
            self.entry7 = ttk.Combobox(self.entryframe, values=problem_types, state='readonly', width=19)
            self.entry7.grid(row=2, column=1, sticky='wns', pady=(5, 5))
        elif str(self.entry5.get()) == '0':
            self.entry6.config(text='-')
            self.entry7 = tk.Button(self.entryframe, text='-', bg='white', bd=0)
            self.entry7.grid(row=2, column=1, sticky='ewns')

    def set_config2(self, event):
        print(self.button3.get())
        if self.button3.get() == 'Part':
            self.configuration = 'Part'
            self.frame.tkraise()
            self.button1.current(0)
        if self.button3.get() == 'Quadrant':
            self.configuration = 'Quadrant'
            self.frame2.tkraise()
            self.button1.current(1)

    def set_quadrant(self, event):
        self.selected_quadrant = self.button4.get()
        if self.selected_quadrant == 'Selecciona un cuadrante':
            return
        if self.plane_type in ['v1000', 'RPB']:
            if self.selected_quadrant == 'Left':
                self.selected_quadrant = 'Izquierdo'
            elif self.selected_quadrant == 'Right':
                self.selected_quadrant = 'Derecho'
            elif self.selected_quadrant == 'Upper':
                self.selected_quadrant = 'Superior'
            elif self.selected_quadrant == 'Lower':
                self.selected_quadrant = 'Inferior'
        # Lista con puntos que unicamente cumplen todos los criterios de herramienta, cuadrante etc (color rojo)
        self.selected_points = []
        for point in self.all_points:
            if any(x in point.process_f_name for x in self.matchers):
                if point.tool_1 in self.selected_tool_types_1:
                    if point.tool in self.selected_tool_types:
                        if point.quadrant == self.selected_quadrant:
                            self.selected_points.append(point)

        # Lista con todos los puntos del cuadrante
        self.quadrant_points = []
        for point in self.all_points:
            if point.quadrant == self.selected_quadrant:
                self.quadrant_points.append(point)

        self.X_quadrant, self.Y_quadrant, self.Z_quadrant = self.get_cords(
            self.quadrant_points)  # Puntos que se ven de color negro

        self.X_selected, self.Y_selected, self.Z_selected = self.get_cords(
            self.selected_points)  # Puntos que se ven de color rojo

        # Hacer plot de cuadrante
        self.plot_quadrant()

    def set_planetype(self, value):
        self.plane_type = value

    def set_part(self):
        index = self.list1.curselection()
        self.selected_part = self.list1.get(index[0])
        print(self.selected_part)

    def change_window(self, window):
        window.tkraise()

    def change_state(self, value):
        self.state = value

    def update_parts_list(self, list):
        self.list1.delete(0, tk.END)
        for name in list:
            self.list1.insert(tk.END, name)

    def update_files_list(self, list):
        self.list1.delete(0, tk.END)
        for name in list:
            self.list1.insert(tk.END, name)

    def check_part_name(self, event):
        text = self.entry1.get()
        if text == '':
            data = self.part_names
        else:
            data = []
            for name in self.part_names:
                if text.lower() in name.lower():
                    data.append(name)
        self.update_parts_list(data)

    def check_file_name(self, event):
        text = self.entry1.get()
        if text == '':
            data = self.files_list
        else:
            data = []
            for name in self.files_list:
                if text.lower() in name.lower():
                    data.append(name)
        self.update_files_list(data)

    def set_active_record(self):
        index = self.list1.curselection()
        selected_record_name = self.list1.get(index[0])
        for file in self.all_files_info:
            if selected_record_name == file.name:
                self.selected_record = file
                self.plane_type = self.selected_record.type
                print(file.name, file.type)

    def set_plane_type(self, event):
        self.plane_type = self.button7.get()

    def create_record(self):
        text = self.entry3.get()
        if text != '' and self.plane_type in ['v900', 'v1000', 'RPB']:
            self.change_state('LOADING')
            # Anadir registro a la lista de registros creados
            self.add_record_info(self.entry3.get(), self.plane_type)
            # Crear archivo excel
            main_root = os.path.dirname(os.path.abspath(__file__))
            f = os.path.join(main_root, 'Records')
            file_root = os.path.join(f, 'Empty_record.xlsx')
            print(file_root)
            excel = load_workbook(filename=file_root)
            sheet = excel['Sheet1']
            new_file_root = os.path.join(f, str(self.selected_record.name) + '.xlsx')
            print(new_file_root)
            excel.save(new_file_root)
            # Cerrar pantalla de creacion de registro
            self.create_record_screen.destroy()

    def add_entry(self):
        if str(self.entry5.get()) == '1':
            point = RegisteredPoint(self.entry4['text'], self.entry5.get(),
                                    self.entry6['text'], self.entry7.get(), self.entry8['text'])
        elif str(self.entry5.get()) == '0':
            point = RegisteredPoint(self.entry4['text'], self.entry5.get(),
                                    self.entry6['text'], self.entry7['text'], self.entry8['text'])
        if point.part_name != '' and point.quality != '' and point.id != '' and point.problem_type != '':
            for registered_point in self.registered_points:
                if point.id == '-' and point.part_name == registered_point.part_name:
                    tk.messagebox.showerror(title='Error al registrar', message='La parte ya fue registrada')
                    return
                if point.id == registered_point.id and point.id != '-':
                    tk.messagebox.showerror(title='Error al registrar', message='El punto ya fue registrado')
                    return
                if point.part_name == registered_point.part_name and registered_point.id == '-':
                    tk.messagebox.showerror(title='Error al registrar', message='La parte ya fue registrada')
                    return
            self.table.insert(parent='', index=self.registered_point_table_id,
                              iid=self.registered_point_table_id,
                              values=(point.part_name, point.quality, point.id, point.problem_type, point.quadrant))
            self.registered_point_table_id += 1
            self.update_registry(point)
            self.registered_points.append(point)
            self.update_parts_list_color()

    def delete_entry(self):
        def set_to_delete_part_ids(event):
            part_name = button.get()
            ids = []
            for point in self.registered_points:
                if point.part_name == part_name:
                    ids.append(point.id)
            button2.config(values=ids)

        def check_input(event):
            value = button.get()
            if value == '':
                button.config(values=part_names)
            else:
                data = []
                for item in part_names:
                    if value.lower() in item.lower():
                        data.append(item)
                button.config(values=data)

        def delete_selected_entry():
            part_name = str(button.get())
            id = button2.get()
            try:
                id = int(id)
            except:
                pass
            if part_name not in part_names:
                tk.messagebox.showerror(title='Error al eliminar', message='Introduce un nombre de parte valido')
                return
            point_to_del = None
            row = 2
            row_number = None
            index = 0
            for point in self.registered_points:
                if point.part_name == part_name and point.id == id:
                    row_number = row
                    self.registered_points.pop(index)
                    break
                elif point.part_name == part_name and point.id == str(id):
                    row_number = row
                    self.registered_points.pop(index)
                    break
                row = row + 1
                index = index + 1
            color_index = self.part_names.index(part_name)
            self.list1.itemconfig(color_index, bg='#CBCBCB')
            main_root = os.path.dirname(os.path.abspath(__file__))
            f = os.path.join(main_root, 'Records')
            file_root = os.path.join(f, str(self.selected_record.name) + '.xlsx')
            excel = load_workbook(filename=file_root)
            sheet = excel['Sheet1']
            sheet.delete_rows(row_number, 1)
            excel.save(file_root)
            self.add_table_data(self.registered_points)
            delete_entry_screen.destroy()

        # Start TKinter
        delete_entry_screen = tk.Tk()
        # Create a style
        style = ttk.Style(delete_entry_screen)
        delete_entry_screen.tk.call('source', 'azure.tcl')
        style.theme_use('azure')

        # Screen settings
        windowWidth = 300
        windowHeight = 180
        screenWidth = delete_entry_screen.winfo_screenwidth()
        screenHeight = delete_entry_screen.winfo_screenheight()
        xCordinate = int((screenWidth / 2) - (windowWidth / 2))
        yCordinate = int((screenHeight / 2) - (windowHeight / 2))
        delete_entry_screen.geometry("{}x{}+{}+{}".format(windowWidth, windowHeight, xCordinate, yCordinate))
        part_names = []
        ids = []
        for point in self.registered_points:
            part_names.append(point.part_name)
        # Frame
        frame = tk.LabelFrame(delete_entry_screen, text='Eliminar un dato del registro')
        frame.pack(fill='both', expand=True, padx=15, pady=15)
        n_cols = 2
        n_rows = 3
        weight = 1
        for i in range(0, n_cols):
            tk.Grid.columnconfigure(frame, i, weight=weight)
        for i in range(0, n_rows):
            tk.Grid.rowconfigure(frame, i, weight=weight)
        # Label de nombre de parte
        label = tk.Label(frame, text='Part Name:', anchor='w')
        label.grid(row=0, column=0, sticky='w')
        # Boton de parte
        button = ttk.Combobox(frame, values=part_names, width=22)
        button.grid(row=0, column=1, sticky='w', pady=5)
        button.bind("<<ComboboxSelected>>", set_to_delete_part_ids)
        button.bind('<KeyRelease>', check_input)
        # Label de ID
        label2 = tk.Label(frame, text='ID:', anchor='w')
        label2.grid(row=1, column=0, sticky='w')
        # Boton de id
        button2 = ttk.Combobox(frame, values=ids, width=10, state='readonly')
        button2.grid(row=1, column=1, sticky='w')

        # Boton de borrar
        button3 = ttk.Button(frame, text='Delete', style='AccentButton', command=lambda: delete_selected_entry())
        button3.grid(row=2, column=0, pady=5, columnspan=2)

    def update_parts_list_color(self):
        for point in self.registered_points:
            index = self.part_names.index(point.part_name)
            self.list1.itemconfig(index, bg='#737373')

    # Methods
    def on_pick_point(self, event):
        thisline = event.artist
        xdata = thisline.get_xdata()
        ydata = thisline.get_ydata()
        ind = event.ind
        xvalue = np.take(xdata, ind)
        yvalue = np.take(ydata, ind)
        if self.selected_quadrant == 'Upper' or self.selected_quadrant == 'Lower' or self.selected_quadrant == 'Superior' or self.selected_quadrant == 'Inferior':
            if xvalue[0] in self.X_selected and yvalue[0] in self.Y_selected:
                self.clicked_point = self.find_pointXY(self.selected_points, xvalue[0], yvalue[0])
        if self.selected_quadrant == 'Right' or self.selected_quadrant == 'Left' or self.selected_quadrant == 'Derecho' or self.selected_quadrant == 'Izquierdo':
            if xvalue[0] in self.X_selected and yvalue[0] in self.Z_selected:
                self.clicked_point = self.find_pointXZ(self.selected_points, xvalue[0], yvalue[0])
        self.point_info_text = "Location: ({},{})\nName: {} \nID: {} \nParts: \n{}".format(
            xvalue[0], yvalue[0],
            self.clicked_point.process_f_name,
            self.clicked_point.id,
            self.clicked_point.part_names)
        for part_name in self.clicked_point.part_names:
            print(part_name)
            if part_name != '-':
                self.plot_partANDpoint(self.all_parts, part_name, self.clicked_point)

    def get_cords(self, points_list):
        Xs = []
        Ys = []
        Zs = []
        for point in points_list:
            Xs.append(point.xe)
            Ys.append(point.ye)
            Zs.append(point.ze)

        return Xs, Ys, Zs

    def find_pointXYZ(self, point_list, x, y, z):
        for point in point_list:
            if point.xe == x and point.ye == y and point.ze == z:
                return point

    def find_pointXY(self, point_list, x, y):
        for point in point_list:
            if point.xe == x and point.ye == y:
                return point

    def find_pointXZ(self, point_list, x, z):
        for point in point_list:
            if point.xe == x and point.ze == z:
                return point

    def plot_partANDpoint(self, all_parts, part_name, clicked_point):
        for part in all_parts:
            if part.name == part_name and part_name != '-':
                print('Coords del punto seleccionado: ({}, {}, {})'.format(clicked_point.xe, clicked_point.ye,
                                                                           clicked_point.ze))
                x = part.x
                y = part.y
                z = part.z
                x.remove(clicked_point.xe)
                y.remove(clicked_point.ye)
                z.remove(clicked_point.ze)
                fig2 = plt.figure(figsize=(5, 4))
                ax_2 = fig2.add_subplot(projection='3d')
                ax_2.scatter(x, y, z, color='blue')
                ax_2.scatter(clicked_point.xe, clicked_point.ye, clicked_point.ze, color='red')
                ax_2.set_title(part_name)

                # Create widget
                miniframe = tk.Frame(self.part_window)
                self.part_window.add(miniframe, text=part_name)
                canvas_plot = FigureCanvasTkAgg(fig2, miniframe)
                canvas_plot.get_tk_widget().pack(fill='both', side='top', expand=True)
                # Caja de informacion
                infobox = tk.LabelFrame(miniframe, text='Datos del punto')
                infobox.pack(fill='both', side='bottom', expand=True)
                infolabel = tk.Label(infobox, text=self.point_info_text, font=('Arial', '14'))
                infolabel.pack(fill='both')

                # Close part button
                button = ttk.Button(miniframe, text='X', style='AccentButton',
                                    command=lambda: [miniframe.destroy(), button.destroy()])
                button.place(relx=.9, rely=0, width=60, height=60)
                break

    def plot_quadrant(self):
        self.fig = plt.figure(dpi=100, frameon=False, figsize=(6, 5))
        self.figure = self.fig.add_subplot()
        self.figure.set_title(self.selected_quadrant)
        if self.selected_quadrant == 'Upper' or self.selected_quadrant == 'Lower' or self.selected_quadrant == 'Superior' or self.selected_quadrant == 'Inferior':
            self.figure.plot(self.X_quadrant, self.Y_quadrant, 'o', markersize=1, color='black')
            self.figure.plot(self.X_selected, self.Y_selected, 'o', markersize=2, color='red', picker=5)
        elif self.selected_quadrant == 'Right' or self.selected_quadrant == 'Left' or self.selected_quadrant == 'Derecho' or self.selected_quadrant == 'Izquierdo':
            self.figure.plot(self.X_quadrant, self.Z_quadrant, 'o', markersize=1, color='black')
            self.figure.plot(self.X_selected, self.Z_selected, 'o', markersize=2, color='red', picker=5)
        # Comando para seleccionar parte
        self.fig.canvas.mpl_connect('pick_event', self.on_pick_point)

        # Crear el widget
        miniframe = tk.LabelFrame(self.frame2, text='Cuadrante')
        miniframe.grid(row=3, column=0, rowspan=2, columnspan=2, sticky='ewns', pady=(0, 20), padx=(10, 10))

        self.canvas_plot = FigureCanvasTkAgg(self.fig, miniframe)
        self.canvas_plot.get_tk_widget().pack(fill='both', side='top', expand=True)
        toolbar = NavigationToolbar2Tk(self.canvas_plot, miniframe).pack(fill='both', side='top', expand=True)

    def plot_part(self, all_parts, part_name):
        for part in all_parts:
            if part.name == part_name and part_name != '-':
                coords_array = pd.DataFrame({'X': part.x, 'Y': part.y, 'Z': part.z}).to_numpy()  # Nparray
                point_names, point_ids, point_parts = self.get_part_info(part)
                self.visualize3DData_Part(coords_array, point_names, point_ids, part, point_parts)
                break

    def visualize3DData_Part(self, A, Name, IDs, part, point_parts):
        """Visualize data in 3d plot with popover next to mouse position"""
        self.fig_2 = plt.figure(figsize=(5, 4))
        self.fig_2ax_2 = self.fig_2.add_subplot(projection='3d')
        self.fig_2ax_2.scatter(part.x, part.y, part.z, color='blue')
        self.fig_2ax_2.set_title(part.name)
        self.selected_registered_point_id = 0
        self.hover_id = 0

        def distance(point, event):
            """Return distance between mouse position and given data point
            Args:
                point (np.array): np.array of shape (3,), with x,y,z in data coords
                event (MouseEvent): mouse event (which contains mouse position in .x and .xdata)
            Returns:
                distance (np.float64): distance (in screen coords) between mouse pos and data point
            """
            assert point.shape == (3,), "distance: point.shape is wrong: %s, must be (3,)" % point.shape

            # Project 3d data space to 2d data space
            x2, y2, _ = proj3d.proj_transform(point[0], point[1], point[2], plt.gca().get_proj())
            # Convert 2d data space to 2d screen space
            x3, y3 = self.fig_2ax_2.transData.transform((x2, y2))

            return np.sqrt((x3 - event.x) ** 2 + (y3 - event.y) ** 2)

        def calcClosestDatapoint(A, event):
            """"Calculate which data point is closest to the mouse position.
            Args:
                A (np.array) - array of points, of shape (numPoints, 3)
                event (MouseEvent) - mouse event (containing mouse position)
            Returns:
                smallestIndex (int) - the index (into the array of points A) of the element closest to the mouse position
            """
            distances = [distance(A[i, 0:3], event) for i in range(A.shape[0])]
            return np.argmin(distances)

        def annotatePlot(A, index, mouse, counter):
            """Create popover label in 3d chart
            Args:
                A (np.array) - array of points, of shape (numPoints, 3)
                index (int) - index (into points array A) of item which should be printed
            Returns:
                None
            """
            # If we have previously displayed another label, remove it first
            if mouse == True:
                if hasattr(annotatePlot, 'label'):
                    if counter > 0:
                        annotatePlot.label.remove()
                # Get data point from array of points X, at position index
                x2, y2, _ = proj3d.proj_transform(A[index, 0], A[index, 1], A[index, 2], self.fig_2ax_2.get_proj())
                annotatePlot.label = plt.annotate(
                    "Name:{}\nID:{}\nPartes:{} ".format(Name[index], IDs[index], point_parts[index][0]),
                    xy=(x2, y2), xytext=(-20, 20), textcoords='offset points', ha='right',
                    va='bottom',
                    bbox=dict(boxstyle='round,pad=0.5', fc='grey', alpha=0.5),
                    arrowprops=dict(arrowstyle='->', connectionstyle='arc3,rad=0'))
                self.fig_2.canvas.draw()
                self.hover_id = IDs[index]
            elif mouse == False:
                if hasattr(annotatePlot, 'label'):
                    annotatePlot.label.remove()
                self.fig_2.canvas.draw()

        def onMouseMotion(event):
            """Event that is triggered when mouse is moved. Shows text annotation over data point closest to mouse."""
            global mouse_in_plot, counter
            closestIndex = calcClosestDatapoint(A, event)
            annotatePlot(A, closestIndex, mouse_in_plot, counter)
            counter = counter + 1

        def mouse_in(event):
            global mouse_in_plot, counter
            mouse_in_plot = True
            counter = 0

        def mouse_out(event):
            global mouse_in_plot, counter
            mouse_in_plot = False
            counter = 0
            closestIndex = calcClosestDatapoint(A, event)
            annotatePlot(A, closestIndex, mouse_in_plot, counter)

        def mouse_click(event):
            self.selected_registered_point_id = self.hover_id
            print(self.selected_registered_point_id)
            self.entry6.config(text=self.selected_registered_point_id)
            self.entry4.config(text=self.selected_part)
            self.entry8.config(text=part.quadrant)

        self.fig_2.canvas.mpl_connect('motion_notify_event', onMouseMotion)  # on mouse motion
        self.fig_2.canvas.mpl_connect('figure_enter_event', mouse_in)
        self.fig_2.canvas.mpl_connect('figure_leave_event', mouse_out)
        self.fig_2.canvas.mpl_connect('button_press_event', mouse_click)
        # Crear el widget
        miniframe = tk.LabelFrame(self.frame, text=part.name)
        miniframe.grid(row=3, column=1, sticky='ewns', pady=(0, 20), padx=(0, 10), rowspan=2)

        self.canvas_part = FigureCanvasTkAgg(self.fig_2, miniframe)
        self.canvas_part.get_tk_widget().pack(fill='both', expand=True)

    def get_part_info(self, part):
        point_names = []
        point_ids = []
        point_parts = []
        for i in range(0, len(part.x)):
            point = self.find_pointXYZ(self.all_points, part.x[i], part.y[i], part.z[i])
            point_names.append(point.process_f_name)
            point_ids.append(point.id)
            parts_list = point.part_names
            string = ''
            while '-' in parts_list:
                parts_list.remove('-')
            for name in parts_list:
                string = string + '{}\n'.format(name)
            list2 = [string]
            point_parts.append(list2)
        return point_names, point_ids, point_parts

    def read_file_names(self):
        main_root = os.path.dirname(os.path.abspath(__file__))
        file_root = os.path.join(main_root, 'Created_records.xlsx')
        excel = pd.read_excel(file_root)
        names = excel['record_name']
        types = excel['type']
        all_files_info = []
        for i in range(len(names)):
            name = names[i]
            type = types[i]
            info = FileInfo(name, type)
            all_files_info.append(info)
        return all_files_info

    def add_record_info(self, name, planetype):
        file = FileInfo(name, planetype)
        self.selected_record = file
        main_root = os.path.dirname(os.path.abspath(__file__))
        file_root = os.path.join(main_root, 'Created_records.xlsx')
        print(file_root)
        excel = load_workbook(filename=file_root)
        sheet = excel['Sheet1']
        new_row = [file.name, file.type]
        sheet.append(new_row)
        excel.save('Created_records.xlsx')

    def read_record(self):
        main_root = os.path.dirname(os.path.abspath(__file__))
        f = os.path.join(main_root, 'Records')
        file_root = os.path.join(f, str(self.selected_record.name) + '.xlsx')
        excel = pd.read_excel(file_root)
        part_names = excel['part_name']
        qualities = excel['quality']
        ids = excel['id']
        problem_types = excel['problem_type']
        quadrants = excel['quadrant']
        registered_points = []
        for i in range(len(part_names)):
            part_name = part_names[i]
            quality = qualities[i]
            id = ids[i]
            problem_type = problem_types[i]
            quadrant = quadrants[i]
            point = RegisteredPoint(part_name, quality, id, problem_type, quadrant)
            registered_points.append(point)
        return registered_points

    def add_table_data(self, registered_points_list):
        self.registered_point_table_id = 0
        self.table.delete(*self.table.get_children())
        for registered_point in registered_points_list:
            part_name = registered_point.part_name
            quality = registered_point.quality
            id = registered_point.id
            problem_type = registered_point.problem_type
            quadrant = registered_point.quadrant
            self.table.insert(parent='', index=self.registered_point_table_id,
                              iid=self.registered_point_table_id, values=(part_name, quality, id, problem_type, quadrant))
            self.registered_point_table_id += 1

    def update_registry(self, point):
        main_root = os.path.dirname(os.path.abspath(__file__))
        f = os.path.join(main_root, 'Records')
        file_root = os.path.join(f, str(self.selected_record.name) + '.xlsx')
        excel = load_workbook(filename=file_root)
        sheet = excel['Sheet1']
        new_row = [point.part_name, point.quality, point.id, point.problem_type, point.quadrant]
        sheet.append(new_row)
        excel.save(file_root)

    def get_point_part_names(self):
        dict_point_parts = {}
        if self.plane_type == 'v900':
            file_name = 'PARTES_INTERES_V900_UNICO.xlsx'
        if self.plane_type == 'v1000':
            file_name = 'PARTES_INTERES_V1000_UNICO.xlsx'
        if self.plane_type == 'RPB':
            file_name = 'PARTES_INTERES_RPB_UNICO.xlsx'
        main_root = os.path.dirname(os.path.abspath(__file__))
        file_root = os.path.join(main_root, file_name)
        df = pd.read_excel(file_root)
        names = df['DESCRIPTION']
        ids = df['ID']
        for i in range(len(ids)):
            idd = ids[i]
            name = names[i]
            if idd not in dict_point_parts:
                dict_point_parts[idd] = [name]
            else:
                dict_point_parts[idd].append(name)
        return dict_point_parts

    def get_all_parts(self):
        if self.plane_type == 'v900':
            file_name = 'PARTES_INTERES_V900_UNICO.xlsx'
        if self.plane_type == 'v1000':
            file_name = 'PARTES_INTERES_V1000_UNICO.xlsx'
        if self.plane_type == 'RPB':
            file_name = 'PARTES_INTERES_RPB_UNICO.xlsx'
        main_root = os.path.dirname(os.path.abspath(__file__))
        file_root = os.path.join(main_root, file_name)
        df = pd.read_excel(file_root)
        names = df['DESCRIPTION']
        part_names = []
        for i in range(len(names)):
            name = names[i]
            if name not in part_names:
                part_names.append(name)
        ids = df['ID']
        xs = df['Xe']
        ys = df['Ye']
        zs = df['Ze']
        quadrants = df['QUADRANT']
        all_parts = []
        for part_name in part_names:
            added_ids = []
            X = []
            Y = []
            Z = []
            quadrant = ''
            for i in range(len(names)):
                name = names[i]
                x = xs[i]
                y = ys[i]
                z = zs[i]
                id = ids[i]
                if id not in added_ids and name == part_name:
                    X.append(x)
                    Y.append(y)
                    Z.append(z)
                    added_ids.append(id)
                    quadrant = quadrants[i]
            p = Part(part_name, X, Y, Z, quadrant)
            print(p.name, p.quadrant)
            all_parts.append(p)
        return part_names, all_parts


app = App()
